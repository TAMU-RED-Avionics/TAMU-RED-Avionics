import os
import re
import math
from collections import deque
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QDialog, QLabel, QDialogButtonBox, QCheckBox, QMessageBox, QGroupBox, QComboBox
from PyQt5.QtCore import QDate, Qt, QTimer, QDateTime
from PyQt5.QtGui import QFont
from PyQt5.QtCore import QObject, pyqtSignal
from GUI_COMMS import EthernetClient

_TEMPLATE_TOKEN_RE = re.compile(r"\{([A-Za-z0-9_]+)\}")


def render_message_template(template: str, context: dict) -> str:
    """Substitute {TOKEN} placeholders in a user-authored abort/warn message.

    `context` maps both special keywords (e.g. "soak_ms", "threshold",
    "value") and sensor names (e.g. "P5") to their current values. Lookup is
    case-sensitive first (so a lowercase keyword like {soak_ms} isn't shadowed
    by an unrelated sensor), then falls back to an upper-cased match so sensor
    tokens work regardless of the case the user typed. Unknown tokens are left
    as literal text (e.g. "{TYPO}") so a mistake is visible rather than silently
    dropped.
    """
    def _repl(m: "re.Match") -> str:
        key = m.group(1)
        value = context.get(key)
        if value is None:
            value = context.get(key.upper())
        if value is None:
            return m.group(0)
        return f"{value:.2f}" if isinstance(value, float) else str(value)

    return _TEMPLATE_TOKEN_RE.sub(_repl, template)


# This may be necessary for ongoing refactors but currently has no use
class Signals(QObject):
    abort_triggered = pyqtSignal(str, str)
    safe_state = pyqtSignal()
    connected = pyqtSignal()
    disconnected = pyqtSignal(str)
    warnings_changed = pyqtSignal(list)
    
    valve_updated = pyqtSignal(str, str)     # can be open, closed, or pending a response from the MCU
    sensor_updated = pyqtSignal(str, float, float)      # sensor_name, val, timestamp
    system_status = pyqtSignal(str)

    # Calculated performance channels (Mdot, Thrust, ...) recomputed whenever
    # a sensor value changes. Payload is {channel_id: value_or_None} plus the
    # two synthesized totals under the reserved keys "__mdot_total__" and
    # "__isp__".
    calc_values_changed = pyqtSignal(dict)

    # Auto-abort countdown signals (for thread-safe dialog creation)
    countdown_start = pyqtSignal(int, str, str)  # initial_seconds, valve_name, cmd_type
    countdown_update = pyqtSignal(int)  # remaining_seconds
    countdown_close = pyqtSignal()


"""
GUI CONTROLLER

This class object is the daddy of the entire GUI. It is intended to manage state. All actions that 
are called by the windows which cant be contained locally are functions within this object. This 
controller will shit out signals depending on both information it retrieves from the EthernetClient
as well as the action functions connected to different buttons.

Think of it as the entire backend managed in one spot.

The rest of the windows are configured to take in this controller as an initializer object. Each
one of them will connect different signals inside here to their internal update actions, and will
connect functions here to their buttons in order to configure connections.

"""
class GUIController:
    def __init__(self, parent: QWidget):
        self.parent = parent

        # Create signals FIRST before EthernetClient (which references them in callbacks)
        self.signals = Signals()
        
        # The EthernetClient will connect to the "flight" MCU and listen for packets in a backend thread
        self.ethernet_client = EthernetClient(
            receive_callback = self.handle_new_data,
            log_event_callback = self.log_event,
            connect_callback = self.handle_connect,
            disconnect_callback = lambda reason: self.signals.disconnected.emit(reason)
        )
        

        # Explaining the disconnect loop - the ethernet client calls the disconnect_callback
        # in a separate thread, therefore we must use a signal that pops out of it and back in here
        # to safely change things in the main thread as a result
        # self.signals.connected.connect(self.handle_connect)
        self.signals.disconnected.connect(self._on_disconnected)
        self.signals.abort_triggered.connect(self.handle_abort)
        self.signals.valve_updated.connect(self.update_valve_state)
        
        # Connect countdown signals for thread-safe dialog creation
        self.signals.countdown_start.connect(self.show_abort_countdown_dialog)
        self.signals.countdown_update.connect(self.update_abort_countdown_dialog)
        self.signals.countdown_close.connect(self.close_abort_countdown_dialog)

        # For file recording (xlsx)
        self._wb: Optional["Workbook"] = None       # openpyxl Workbook
        self._ws = None                              # active Sheet
        self._xlsx_path: str = ""                   # full path of open file

        # Sensor columns - rebuilt from project on load_project(); fallback to empty lists
        self.pt_keys:    list = []   # pressure transducers
        self.tc_keys:    list = []   # thermocouples
        self.lc_keys:    list = []   # load cells

        self.current_batch_sensors: set = set()
        self.latest_teensy_ts = ""

        # Sensor pipeline:  wire value → calibration (gain·x + offset) → tare.
        # raw_sensor_values holds calibrated-but-untared values (tare math
        # works on top of calibration).
        self.raw_sensor_values: dict = {}
        self.tare_offsets: dict = {}
        self.calibrations: dict = {}        # name -> (gain, offset)

        # Health tracking: last receive time + recent uncalibrated samples
        # (used for stale / flatline detection and the calibration wizard)
        self.sensor_last_rx: dict = {}      # name -> unix seconds
        self.sensor_recent: dict = {}       # name -> deque of wire values

        # These are constants and dictionaries that the UI needs to be tracked
        self.lockout = True                     # default to lockout until a connection starts

        #self.ncs3_opened_due_to_p2 = False  # legacy compat
        #self.abort_modes: dict = {}         # legacy compat
        self.pre_abort_valve_states: dict = {}
        self.current_sensor_values: dict = {}
        self.manual_valve_buttons: dict = {}
        self.abort_check_interval = 10  # ms
        self.throttling_enabled = False
        self.gimbaling_enabled = False
        self.manual_valve_dialog: QDialog = None
        self.recording = False

        # Abort countdown dialog tracking
        self.abort_countdown_dialog: Optional[QDialog] = None
        self.abort_countdown_label: Optional[QLabel] = None

        # Project / rule engine state
        self.project = None       # PIDProject set by MainWindow via load_project()
        self.project_path = None  # str path to .red file

        # Rule engine runtime state
        self._rule_violation_start: dict = {}  # rule_id -> ms timestamp when violation started
        self._rule_valve_open: dict = {}       # rule_id -> bool, tracks auto-opened valves
        self._sensor_rules: list = []          # compiled from component thresholds
        self._logic_rules: list = []           # AbortRule expression type
        self._active_warnings: dict[str, str] = {}

        # Valve Command Reliability
        self.pending_valve_commands = {} # valve_name -> {"state": bool, "last_sent": timestamp, "retries": int}
        self.valve_retry_timeout = 200 # ms
        self.valve_max_retries = 5
        self.retry_timer = QTimer()
        self.retry_timer.timeout.connect(self.check_valve_command_timeouts)
        self.retry_timer.start(100) # Check every 100ms

        # Initial valve states: False = closed (red), True = open (green)
        self.valve_states: dict[str, bool] = {
            "NCS1": False,
            "NCS2": False,
            "NCS3": False,
            "NCS5": False,
            "NCS6": False,
            "LA-BV1": False,
            "GV-1": False,
            "GV-2": False
        }

        # hw_ids that are igniters, not valves - kept out of any automated
        # cycling (health-check valve checkout). They still live in
        # valve_states (so "Fully Closed" can safe one left firing).
        self.igniter_hw_ids: set = set()

        # Sequence execution runtime state
        self._sequence_timers: list[QTimer] = []   # kept alive during a fire sequence run

        # Calculated performance channels (Mdot, Thrust, Isp, ...) - recomputed
        # any time a sensor value updates.
        self.calc_values: dict = {}
        self.signals.sensor_updated.connect(lambda *_args: self._recompute_calc_channels())

        self.setup_abort_monitor()
    

    def _summarize_connection_error(self, reason: str) -> str:
        """Map verbose socket diagnostics to concise UI status text."""
        msg = (reason or "").lower()
        if "timed out" in msg:
            return "Connection Timeout"
        if "already in use" in msg:
            return "Port In Use"
        if "network unreachable" in msg:
            return "Network Unreachable"
        if "host" in msg and "unreachable" in msg:
            return "Host Unreachable"
        if "permission denied" in msg:
            return "Permission Denied"
        return "Connection Failed"
    
    def _on_disconnected(self, reason: str):
        """A dropped link is treated as an abort; an operator-requested
        disconnect (graceful_disconnect) just returns to the idle state."""
        if self.ethernet_client.intentional_disconnect:
            self.ethernet_client.intentional_disconnect = False
            self.lockout = True   # same as pre-connection: nothing commandable
            self.signals.system_status.emit("Disconnected (safe)")
            return
        self.signals.abort_triggered.emit("DISCONNECTED", reason)

    def request_safe_disconnect(self):
        """Operator disconnect: notify the MCU (SFE) and tear the link down
        without the abort path firing."""
        self.log_event("OPERATOR_DISCONNECT", "Safe disconnect requested")
        self.ethernet_client.graceful_disconnect()

    def handle_connect(self, success: bool):
        if success:
            self.ethernet_client.set_system_state("CONNECTED")
            # Clear any stale status text left over from a prior disconnect
            # (e.g. "Disconnected (safe)") so it doesn't linger after reconnect.
            self.signals.system_status.emit("")
            self.signals.connected.emit()
        else:
            detailed_reason = self.ethernet_client.last_connection_error or "Connection failed"
            concise_reason = self._summarize_connection_error(detailed_reason)
            self.signals.disconnected.emit(concise_reason)
    
    def setup_abort_monitor(self):
        self.abort_timer = QTimer()
        self.abort_timer.timeout.connect(self.check_abort_conditions)
        self.abort_timer.start(self.abort_check_interval)

    def load_project(self, project, path: str = None):
        self.project = project
        if path:
            self.project_path = path
        self._rebuild_sensor_keys(project)
        self._rebuild_valve_states(project)
        self.load_calibrations_from_project(project)
        self.reload_abort_rules(project)
        self.ethernet_client.load_project_maps(project)
        self._emit_warnings()
        self._recompute_calc_channels()

    def _rebuild_valve_states(self, project):
        """Rebuild valve_states from project valve components (hw_id -> bool).
        Existing states are preserved for valves that carry over; new valves start closed."""
        from PID_SCHEMA import (
            COMP_VALVE, COMP_THROTTLE_VALVE, COMP_BALL_VALVE, COMP_GLOBE_VALVE,
            COMP_SOLENOID, COMP_IGNITER,
        )
        VALVE_TYPES = (COMP_VALVE, COMP_THROTTLE_VALVE, COMP_BALL_VALVE,
                       COMP_GLOBE_VALVE, COMP_SOLENOID, COMP_IGNITER)

        if not project:
            return

        new_states: dict[str, bool] = {}
        new_igniters: set = set()
        for comp in project.components.values():
            if comp.type not in VALVE_TYPES:
                continue
            hw_id = comp.extras.get("hw_id") or comp.label
            if not hw_id:
                continue
            # Preserve state if valve was already known, default to closed
            new_states[hw_id] = self.valve_states.get(hw_id, False)
            if comp.type == COMP_IGNITER:
                new_igniters.add(hw_id)

        self.valve_states = new_states
        self.igniter_hw_ids = new_igniters

    def _rebuild_sensor_keys(self, project):
        from PID_SCHEMA import COMP_PRESSURE, COMP_TEMPERATURE, COMP_LOAD_CELL
        pt, tc, lc = [], [], []
        if project:
            for comp in project.components.values():
                hw_id = comp.extras.get("hw_id", comp.label) or comp.label
                if not hw_id:
                    continue
                if comp.type == COMP_PRESSURE:
                    pt.append(hw_id)
                elif comp.type == COMP_TEMPERATURE:
                    tc.append(hw_id)
                elif comp.type == COMP_LOAD_CELL:
                    lc.append(hw_id)
        self.pt_keys = sorted(set(pt), key=lambda x: (len(x), x))
        self.tc_keys = sorted(set(tc), key=lambda x: (len(x), x))
        self.lc_keys = sorted(set(lc), key=lambda x: (len(x), x))

    def _generate_filename(self) -> str:
        """
        Auto-generate a filename following the convention:
            {ProjectName}{Version}_{MMDDYYYY}_test{N}.xlsx
        The test number increments so existing files are never overwritten.
        If no project is loaded, falls back to 'DAQ_{date}_test{N}.xlsx'.
        """
        if self.project:
            raw_version = getattr(self.project, "version", "") or ""
            proj_part = f"{self.project.name}{raw_version}"
        else:
            proj_part = "DAQ"

        date_part = datetime.now().strftime("%m%d%Y")
        base = f"{proj_part}_{date_part}"

        data_dir = "Data"
        os.makedirs(data_dir, exist_ok=True)

        n = 1
        while True:
            name = f"{base}_test{n}.xlsx"
            if not os.path.exists(os.path.join(data_dir, name)):
                return name
            n += 1

    def reload_abort_rules(self, project):
        self._sensor_rules = []
        self._logic_rules  = []
        self._rule_violation_start.clear()
        self._rule_valve_open.clear()
        self._active_warnings.clear()

        if project is None:
            return

        from PID_SCHEMA import (
            COMP_PRESSURE, COMP_TEMPERATURE, COMP_LOAD_CELL,
            SensorThresholds,
        )
        SENSOR_TYPES = (COMP_PRESSURE, COMP_TEMPERATURE, COMP_LOAD_CELL)

        for cid, comp in project.components.items():
            if comp.type not in SENSOR_TYPES:
                continue

            th_dict = comp.extras.get("thresholds")
            if not th_dict:
                continue

            th = SensorThresholds.from_dict(th_dict)
            hw_id = comp.extras.get("hw_id", comp.label) or cid

            bands = [
                ("mawp",    th.mawp,    th.mawp_action,   None,                                                          None),
                ("meop",    th.meop,    th.meop_action,   th_dict.get("meop_target",   th.target_valve),                  th_dict.get("meop_close_below",   th.close_valve_below)),
                ("relief",  th.relief,  th.relief_action, th_dict.get("relief_target", th.target_valve),                  th_dict.get("relief_close_below", th.close_valve_below)),
            ]

            for band, threshold, action, target, close_below in bands:
                if threshold is None:
                    continue
                rule_id = f"{cid}__{band}"
                self._sensor_rules.append({
                    "id":           rule_id,
                    "hw_id":        hw_id,
                    "band":         band,
                    "threshold":    threshold,
                    "action":       action,
                    "target_valve": target or "",
                    "close_below":  close_below,
                    "soak_ms":      th.soak_ms,
                    "above":        True,
                    # Optional custom {TOKEN} message template for this band -
                    # blank means fall back to the built-in default wording.
                    "message":      th_dict.get(f"{band}_message", ""),
                })

        for rule in project.rules:
            if rule.condition_type == "expression" and rule.enabled and rule.expression.strip():
                self._logic_rules.append(rule)

        self._emit_warnings()

    def _emit_warnings(self):
        messages = sorted(self._active_warnings.values())
        self.signals.warnings_changed.emit(messages)

    def _set_warning(self, warning_id: str, message: str):
        if self._active_warnings.get(warning_id) == message:
            return
        self._active_warnings[warning_id] = message
        self._emit_warnings()

    def _clear_warning(self, warning_id: str):
        if warning_id in self._active_warnings:
            del self._active_warnings[warning_id]
            self._emit_warnings()

    def get_sensor_status(self, hw_name: str, value: float) -> str:
        mawp_val   = None
        meop_val   = None
        relief_val = None
        for sr in self._sensor_rules:
            if sr["hw_id"].upper() != hw_name.upper():
                continue
            if sr["band"] == "mawp":
                mawp_val = sr["threshold"]
            elif sr["band"] == "meop":
                meop_val = sr["threshold"]
            elif sr["band"] == "relief":
                relief_val = sr["threshold"]

        if mawp_val   is not None and value >= mawp_val:
            return "RED"
        if meop_val   is not None and value >= meop_val:
            return "ORANGE"
        if relief_val is not None and value >= relief_val:
            return "BLUE"
        return "DEFAULT"

    # ── Calculated performance channels (Mdot / Thrust / Isp) ──────────────
    # Standard gravity, used for Isp(s) = Thrust(lbf) / (Mdot(lbm/s) * G0).
    # Exposed inside channel expressions too (as G0) for anyone writing a
    # custom formula that needs it.
    G0 = 32.174

    _CALC_SAFE_FUNCS = {
        "sqrt": math.sqrt, "abs": abs, "min": min, "max": max, "pow": pow,
        "G0": G0,
    }

    def _recompute_calc_channels(self):
        """Evaluate every CalcChannel against the current sensor values and
        each channel's own constants, the same Python-safe-expression
        approach already used for abort-rule expressions. Also synthesizes
        Total Mdot (sum of mdot_fuel + mdot_ox roles) and Isp
        (thrust / (total_mdot * G0))."""
        values: dict = {}
        channels = list(getattr(self.project, "calc_channels", []) or []) if self.project else []

        for ch in channels:
            if not ch.enabled or not ch.expression.strip():
                values[ch.id] = None
                continue
            ctx = dict(self._CALC_SAFE_FUNCS)
            ctx.update(self.current_sensor_values)
            ctx.update(ch.constants)
            try:
                values[ch.id] = float(eval(ch.expression, {"__builtins__": {}}, ctx))
            except Exception:
                values[ch.id] = None

        mdot_total = 0.0
        have_mdot = False
        thrust_total = 0.0
        have_thrust = False
        for ch in channels:
            v = values.get(ch.id)
            if v is None:
                continue
            if ch.role in ("mdot_fuel", "mdot_ox"):
                mdot_total += v
                have_mdot = True
            elif ch.role == "thrust":
                thrust_total += v
                have_thrust = True

        values["__mdot_total__"] = mdot_total if have_mdot else None
        values["__thrust_total__"] = thrust_total if have_thrust else None
        if have_thrust and have_mdot and mdot_total > 0:
            values["__isp__"] = thrust_total / (mdot_total * self.G0)
        else:
            values["__isp__"] = None

        self.calc_values = values
        self.signals.calc_values_changed.emit(values)


    def check_abort_conditions(self):
        if not self.current_sensor_values:
            return

        locked_out = self.lockout

        current_time = QDateTime.currentMSecsSinceEpoch()

        self._check_system_pressure_limits(locked_out)

        for sr in self._sensor_rules:
            hw_id = sr["hw_id"].upper()
            value = self.current_sensor_values.get(hw_id)
            if value is None:
                self._rule_violation_start.pop(sr["id"], None)
                continue

            threshold = sr["threshold"]
            violated  = (value >= threshold) if sr["above"] else (value <= threshold)

            if not violated:
                # Clear soak timer
                self._rule_violation_start.pop(sr["id"], None)
                self._clear_warning(sr["id"])
                # Close auto-opened valve if below hysteresis close threshold
                if sr["action"] == "open_valve" and self._rule_valve_open.get(sr["id"]):
                    close_below = sr["close_below"]
                    if close_below is not None and value < close_below:
                        if sr["target_valve"] and sr["target_valve"] in self.valve_states:
                            self.toggle_valve(sr["target_valve"], False)
                        self._rule_valve_open[sr["id"]] = False
                continue

            # Violation detected - handle soak time
            soak = sr["soak_ms"]
            if soak > 0:
                if sr["id"] not in self._rule_violation_start:
                    self._rule_violation_start[sr["id"]] = current_time
                    continue
                elif current_time - self._rule_violation_start[sr["id"]] < soak:
                    continue

            self._fire_sensor_rule(sr, hw_id, value, locked_out)

        for rule in self._logic_rules:
            self._check_logic_rule(rule, current_time, locked_out)

    # Per-sensor threshold band -> display word. Internal dict keys/extras
    # stay "mawp" for backward file compatibility; only the shown text changed.
    _BAND_DISPLAY = {"mawp": "HIGH", "meop": "MEOP", "relief": "RELIEF"}

    def _check_system_pressure_limits(self, locked_out: bool):
        """System-wide backstop across all PTs, independent of (and in
        addition to) the per-sensor thresholds: warn above system MEOP,
        abort at 95% of system MAWP."""
        params = getattr(self.project, "parameters", None) if self.project else None
        if not params:
            return

        mawp = getattr(params, "system_mawp", None)
        meop = getattr(params, "system_meop", None)
        if mawp is None and meop is None:
            return

        abort_limit = 0.95 * mawp if mawp is not None else None

        for hw_id in self.pt_keys:
            key = hw_id.upper()
            value = self.current_sensor_values.get(key)
            if value is None:
                continue

            if abort_limit is not None and value >= abort_limit:
                self._clear_warning(f"system_meop__{key}")
                if not locked_out:
                    ctx = self._sensor_message_context(value=value, threshold=abort_limit,
                                                       limit=abort_limit, mawp=mawp)
                    template = (params.system_mawp_message.strip() if params.system_mawp_message
                               else f"{key} ({{{key}}}) exceeds 95% of system MAWP "
                                    "({mawp} psi -> limit {limit} psi)")
                    detail = render_message_template(template, ctx)
                    self.signals.abort_triggered.emit("system_mawp", detail)
                continue

            if meop is not None and value >= meop:
                ctx = self._sensor_message_context(value=value, threshold=meop, limit=meop)
                template = (params.system_meop_message.strip() if params.system_meop_message
                           else f"{key} ({{{key}}}) is above system MEOP (limit {{limit}} psi)")
                self._set_warning(f"system_meop__{key}", render_message_template(template, ctx))
            else:
                self._clear_warning(f"system_meop__{key}")

    def _sensor_message_context(self, **extra) -> dict:
        """Base context for message templates: every live sensor value plus
        whatever special keywords (value/threshold/soak_ms/...) the caller
        supplies for this specific rule."""
        ctx = dict(self.current_sensor_values)
        ctx.update(extra)
        return ctx

    def _fire_sensor_rule(self, sr: dict, hw_id: str, value: float, locked_out: bool = False):
        action  = sr["action"]
        band    = sr["band"]
        rule_id = sr["id"]
        band_disp = self._BAND_DISPLAY.get(band, band.upper())

        ctx = self._sensor_message_context(
            value=value, threshold=sr["threshold"], soak_ms=sr["soak_ms"],
            hw_id=hw_id, band=band_disp,
        )
        template = sr["message"].strip() if sr.get("message") else (
            "{hw_id} {band} threshold exceeded: {value} (limit {threshold})")
        message = render_message_template(template, ctx)

        if action == "abort":
            if locked_out:
                return
            # Show "P1 HIGH" (derived from the sensor's own hw_id/band), never
            # the internal rule_id ("PT_1__mawp") - nothing here is hardcoded,
            # it's built from this sensor's own configuration.
            self.signals.abort_triggered.emit(f"{hw_id} {band_disp}", message)

        elif action == "warn":
            # Keep a persistent UI warning without escalating to abort.
            self._set_warning(rule_id, message)
            self.log_event("WARN", message)
            self.signals.system_status.emit(f"WARNING: {message}")

        elif action == "open_valve":
            if locked_out:
                return
            target = sr["target_valve"]
            if target and not self._rule_valve_open.get(rule_id):
                if target in self.valve_states:
                    self.toggle_valve(target, True)
                self._rule_valve_open[rule_id] = True
                self.log_event("AUTO_VALVE_OPEN",
                               f"Rule {rule_id}: opened {target} ({hw_id}={value:.2f})")

    def _check_logic_rule(self, rule, current_time: int, locked_out: bool = False):
        try:
            result = bool(eval(rule.expression, {"__builtins__": {}},
                               self.current_sensor_values))
        except Exception:
            return

        rule_id = rule.id
        if not result:
            self._rule_violation_start.pop(rule_id, None)
            self._clear_warning(rule_id)
            # Close valve if hysteresis met
            if rule.action == "open_valve" and self._rule_valve_open.get(rule_id):
                if rule.close_valve_below is not None:
                    # Can't check a sensor value without knowing which one for generic expressions;
                    # use the inverse of the expression result as the close condition instead.
                    self._rule_valve_open[rule_id] = False
                    if rule.target_valve and rule.target_valve in self.valve_states:
                        self.toggle_valve(rule.target_valve, False)
            return

        soak = rule.soak_ms
        if soak > 0:
            if rule_id not in self._rule_violation_start:
                self._rule_violation_start[rule_id] = current_time
                return
            elif current_time - self._rule_violation_start[rule_id] < soak:
                return

        if rule.action == "abort":
            if locked_out:
                return
            message = self._render_logic_rule_message(rule)
            # Show the rule's own description ("high chamber pressure"), set
            # by the operator in Abort Config - never the auto-generated
            # rule_id ("logic_rule_3"). Falls back to the expression, then
            # the id, only if the operator left description blank.
            label = rule.description.strip() or rule.expression.strip() or rule_id
            self.signals.abort_triggered.emit(label, message)
        elif rule.action == "open_valve":
            if locked_out:
                return
            if not self._rule_valve_open.get(rule_id):
                if rule.target_valve and rule.target_valve in self.valve_states:
                    self.toggle_valve(rule.target_valve, True)
                self._rule_valve_open[rule_id] = True
                self.log_event("AUTO_VALVE_OPEN",
                               f"Rule {rule_id}: opened {rule.target_valve}")
        elif rule.action == "warn":
            msg = self._render_logic_rule_message(rule)
            self._set_warning(rule_id, msg)
            self.log_event("WARN", f"Rule {rule_id}: {msg}")

    def _auto_template_from_expression(self, expression: str) -> str:
        """Best-effort default template for a logic rule with no custom
        message: wraps each referenced sensor name with its live value, e.g.
        'P5 > P3' -> 'P5 ({P5}) > P3 ({P3})'."""
        def _repl(m: "re.Match") -> str:
            name = m.group(0)
            return f"{name} ({{{name}}})" if name in self.current_sensor_values else name
        return re.sub(r"[A-Za-z_][A-Za-z0-9_]*", _repl, expression)

    def _render_logic_rule_message(self, rule) -> str:
        """Custom {TOKEN} message_template if the user configured one in Abort
        Config, else an auto-generated 'P5 (600) > P3 (595) for 150ms'-style
        message from the expression - never a blanket dump of every sensor."""
        ctx = self._sensor_message_context(soak_ms=rule.soak_ms, expression=rule.expression)
        template = rule.message_template.strip()
        if not template:
            template = rule.description.strip() or self._auto_template_from_expression(rule.expression)
            if rule.soak_ms > 0 and "{soak_ms}" not in template:
                template += " for {soak_ms}ms"
        return render_message_template(template, ctx)

    def trigger_manual_abort(self):
        """Manual abort button handler (Req 11)"""
        self.signals.abort_triggered.emit(
            "manual_abort",
            "Operator triggered manual abort"
        )

    def show_abort_control(self):
        """Redirects to the Abort Config tab in the main window if possible."""
        # The AbortConfigPage is now a full tab, so just log / no-op.
        # If called from a button in DAQWindow, we emit a signal or show a message.
        QMessageBox.information(
            self.parent,
            "Abort Configuration",
            "Use the 'Abort Config' tab to configure abort thresholds and rules."
        )

    def toggle_abort_mode(self, mode, state):
        """Legacy stub - abort modes are now stored in project rules."""
        pass


    def handle_abort(self, abort_type, reason):
        """Handle abort sequence (Req 11, 20-24)"""
        if self.lockout:
            return

        # Stop any running auto-fire sequence immediately
        self._cancel_sequence()

        self.ethernet_client.set_system_state("ABORT")

        self.pre_abort_valve_states = self.valve_states.copy()

        # Disable all valves
        for valve in self.valve_states.keys():
            self.toggle_valve(valve, False)

        # `reason` is already a specific, relevant message by this point -
        # sensor-threshold and logic-rule aborts render it from a (customizable,
        # see Abort Config) {TOKEN} template naming exactly the sensors
        # involved, e.g. "P5 (600.00) > P3 (595.00) for 150ms". Triggers with
        # no sensor context (manual abort, comms loss, disconnect) just show
        # their reason as-is. Dumping every PT's value on every abort regardless
        # of relevance was more noise than signal, so that's gone.
        QMessageBox.critical(
            self.parent, # has to bind to a real widget
            "ABORT TRIGGERED",
            f"Abort Type: {abort_type}\nReason: {reason}"
        )

        self.lockout = True

        # Log abort event
        self.log_event("ABORT", f"{abort_type}:{reason}")

    def confirm_safe_state(self):
        """Confirm system is safe after abort without any dialog"""
        if self.ethernet_client.connected:
            self.ethernet_client.cancel_auto_abort_countdown()
            self.ethernet_client.set_system_state("SAFE")
            self.lockout = False
            # self.abort_state = False
            self.signals.safe_state.emit()
        
            # Log safe state confirmation
            self.log_event("SAFE_STATE", "Operator confirmed safe state")


    # DAQ RECORDING ------------------------------------------------------------------------------------------------

    def log_event(self, event_type, event_details=""):
        """Log an event to the xlsx file (Req 15)"""
        if "HEARTBEAT" in event_type:
            return

        # Handle auto-abort countdown events - use signals for thread safety
        if event_type.startswith("AUTO_ABORT_COUNTDOWN:"):
            parts = event_type.split(":")
            if len(parts) >= 2:
                action = parts[1]
                if action == "START" and len(parts) >= 3:
                    initial_seconds = int(parts[2])
                    valve_name = parts[3] if len(parts) >= 4 else "UNKNOWN"
                    cmd_type = parts[4] if len(parts) >= 5 else "COMMAND"
                    self.signals.countdown_start.emit(initial_seconds, valve_name, cmd_type)
                elif action == "REMAINING" and len(parts) >= 3:
                    remaining_seconds = int(parts[2])
                    self.signals.countdown_update.emit(remaining_seconds)
                elif action == "CANCELED":
                    self.signals.countdown_close.emit()

        if event_type.startswith("COMMS_AUTO_ABORT:"):
            self.signals.countdown_close.emit()

        if not self._ws:
            return

        self._write_daq_row(teensy_ts=self.latest_teensy_ts, event_type=event_type, event_details=event_details)

    def _write_daq_row(self, teensy_ts="", event_type="", event_details=""):
        """Write one data row to the xlsx worksheet.

        Column layout:
            Local Time (24-hr + date) | Teensy Timestamp | [blank] |
            <pressure sensors...>     | [blank]          |
            <thermocouples...>        | [blank]          |
            <load cells...>           | [blank]          |
            Control State             | Event Type       | Event Details
        """
        if not self._ws:
            return

        row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),  # Local time, 24-hr + date
            teensy_ts,
            "",
        ]

        for key in self.pt_keys:
            val = self.current_sensor_values.get(key.upper(), "")
            row.append(round(val, 4) if isinstance(val, float) else val)
        row.append("")

        for key in self.tc_keys:
            val = self.current_sensor_values.get(key.upper(), "")
            row.append(round(val, 4) if isinstance(val, float) else val)
        row.append("")

        for key in self.lc_keys:
            val = self.current_sensor_values.get(key.upper(), "")
            row.append(round(val, 4) if isinstance(val, float) else val)
        row.append("")

        current_state = getattr(self.ethernet_client, "system_state", "UNKNOWN")
        row.extend([current_state, event_type, event_details])

        self._ws.append(row)

    def get_next_filename(self) -> str:
        return self._generate_filename()

    def start_recording(self) -> bool:
        if not self.project:
            QMessageBox.warning(self.parent, "No Project",
                "Please select a project from the dropdown before recording."
                "The project defines which sensors to log.")
            return False

        filename = self._generate_filename()
        data_dir = "Data"
        os.makedirs(data_dir, exist_ok=True)
        file_path = os.path.join(data_dir, filename)

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "DAQ"

            # ---- Header row ----
            header = [
                "Local Time (YYYY-MM-DD HH:mm:ss)",
                "Teensy Timestamp (µs from boot)",
                "",
            ]
            header += self.pt_keys + [""]
            header += self.tc_keys + [""]
            header += self.lc_keys + [""]
            header += ["Control State", "Event Type", "Event Details"]

            ws.append(header)

            # Style: bold + light-grey fill on header row
            header_fill = PatternFill("solid", fgColor="D9D9D9")
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            self._wb = wb
            self._ws = ws
            self._xlsx_path = file_path

            # Clear batch tracker
            self.current_batch_sensors.clear()

            self.log_event("RECORDING:START")
            self.recording = True

            return True

        except Exception as e:
            QMessageBox.critical(self.parent, "Error", f"Failed to create file: {str(e)}")
            return False

    def stop_recording(self):
        if self._wb:
            self.log_event("RECORDING:STOP")
            try:
                self._wb.save(self._xlsx_path)
            except Exception as e:
                QMessageBox.warning(self.parent, "Save Warning", f"Could not save file:\n{str(e)}")
            self._wb = None
            self._ws = None
            self._xlsx_path = ""
            self.recording = False

    def handle_new_data(self, data_str: str):
        """ Parse teensy timestamp (first token) (Req 4) """
        timestamp = QDateTime.currentMSecsSinceEpoch() / 1000.0     # seconds, but at a precision of 1 ms
        
        parts = data_str.split(sep=",",maxsplit=1)
        teensy_ts = parts[0] if len(parts) > 1 else ""
        sensor_data = parts[1] if len(parts) > 1 else data_str

        readings = sensor_data.strip().split(sep=",")
        for reading in readings:
            if "ABORTED" in reading:
                if "ABORTED:COMMS:" in reading:
                    reason_part = reading.split("ABORTED:COMMS:", 1)[1] if len(reading.split("ABORTED:COMMS:")) > 1 else "Unknown reason"
                    
                    if reason_part.startswith("countdown_expired:"):
                        reason_part = reason_part.replace("countdown_expired:", "", 1)
                    
                    if "OPEN_" in reason_part or "CLOSE_" in reason_part:
                        parts = reason_part.split("_")
                        if len(parts) >= 2:
                            cmd = parts[0]
                            valve = parts[1].split(":")[0]
                            display_reason = f"Valve {valve} {cmd} command failed (timeout)"
                        else:
                            display_reason = reason_part
                    else:
                        display_reason = reason_part
                    self.signals.abort_triggered.emit("comms_auto_abort", display_reason)
                elif "ABORTED:REMOTE_SFE:" in reading:
                    sfe_reason = reading.split("ABORTED:REMOTE_SFE:", 1)[1] if len(reading.split("ABORTED:REMOTE_SFE:")) > 1 else "unknown packet"
                    self.signals.abort_triggered.emit("remote_safe_mode", f"Received SFE from MCU ({sfe_reason})")
                else:
                    # Engine MCU triggered abort
                    self.signals.abort_triggered.emit("engine_abort", "The engine MCU triggered a local abort")

            # Valve command responses
            elif "VALVE_SUCCESS" in reading:
                
                parts = reading.split(':')
                if len(parts) >= 3:
                    valve_name: str = parts[1]
                    new_state: str = "OPEN" if parts[2] == "1" else "CLOSED"

                    print(f"ACK received for {valve_name}: {reading}")

                    # Clear pending command if it exists
                    if valve_name in self.pending_valve_commands:
                        del self.pending_valve_commands[valve_name]

                    self.signals.valve_updated.emit(valve_name, new_state)
                

            elif "VALVE_FAIL" in reading:
                parts = reading.split(':')
                if len(parts) >= 2:
                    valve_name: str = parts[1]
                    
                    print(f"NAK received for {valve_name}: {reading}")

                    # Clear pending command if it exists
                    if valve_name in self.pending_valve_commands:
                        del self.pending_valve_commands[valve_name]

                    prev_state = self.valve_states[valve_name]
                    self.signals.valve_updated.emit(valve_name, "OPEN" if prev_state else "CLOSED")

            # Sensor data
            elif ':' in reading:
                try:
                    parts = reading.split(':', 1)
                    sensor_name = parts[0].strip().upper()
                    wire = float(parts[1].strip())

                    self.sensor_last_rx[sensor_name] = timestamp
                    recent = self.sensor_recent.get(sensor_name)
                    if recent is None:
                        recent = self.sensor_recent[sensor_name] = deque(maxlen=20)
                    recent.append(wire)

                    gain, offset = self.calibrations.get(sensor_name, (1.0, 0.0))
                    raw = wire * gain + offset
                    self.raw_sensor_values[sensor_name] = raw
                    value = raw - self.tare_offsets.get(sensor_name, 0.0)
                    self.current_sensor_values[sensor_name] = value
                    self.signals.sensor_updated.emit(sensor_name, value, timestamp)
                    
                    # Batch sensor updates
                    if sensor_name in self.current_batch_sensors:
                        # We've seen this sensor in the current batch -> flush the row
                        if self._ws:
                            self._write_daq_row(teensy_ts=self.latest_teensy_ts)
                        self.current_batch_sensors.clear()
                    
                    self.current_batch_sensors.add(sensor_name)
                    
                except ValueError:
                    pass
        
        if teensy_ts:
            self.latest_teensy_ts = teensy_ts

    def tare_sensor(self, sensor_name: str, target: float) -> bool:
        """Tare so the sensor's current raw reading displays as `target`.
        Returns False when no reading has been received yet."""
        name = sensor_name.upper()
        raw = self.raw_sensor_values.get(name)
        if raw is None:
            return False
        self.tare_offsets[name] = raw - target
        self.current_sensor_values[name] = target
        timestamp = QDateTime.currentMSecsSinceEpoch() / 1000.0
        self.signals.sensor_updated.emit(name, target, timestamp)
        self.log_event("TARE", f"{name}:{target:g} (offset {raw - target:+.4f})")
        return True

    def clear_tare(self, sensor_name: str):
        name = sensor_name.upper()
        if name not in self.tare_offsets:
            return
        del self.tare_offsets[name]
        raw = self.raw_sensor_values.get(name)
        if raw is not None:
            self.current_sensor_values[name] = raw
            timestamp = QDateTime.currentMSecsSinceEpoch() / 1000.0
            self.signals.sensor_updated.emit(name, raw, timestamp)
        self.log_event("TARE_CLEARED", name)

    def set_calibration(self, sensor_name: str, gain: float, offset: float):
        """Install a linear calibration (display = gain·wire + offset) and
        refresh the sensor's displayed value from its latest wire sample."""
        name = sensor_name.upper()
        self.calibrations[name] = (gain, offset)
        recent = self.sensor_recent.get(name)
        if recent:
            raw = recent[-1] * gain + offset
            self.raw_sensor_values[name] = raw
            value = raw - self.tare_offsets.get(name, 0.0)
            self.current_sensor_values[name] = value
            self.signals.sensor_updated.emit(
                name, value, QDateTime.currentMSecsSinceEpoch() / 1000.0)
        self.log_event("CALIBRATION", f"{name}: gain={gain:.6g} offset={offset:.6g}")

    def clear_calibrations(self):
        self.calibrations.clear()
        self.log_event("CALIBRATION", "All calibrations cleared")

    def load_calibrations_from_project(self, project):
        """Pick up per-sensor cal_gain / cal_offset stored in component extras."""
        if not project:
            return
        for comp in project.components.values():
            hw_id = (comp.extras.get("hw_id") or comp.label or comp.id)
            try:
                gain = float(comp.extras["cal_gain"])
                offset = float(comp.extras.get("cal_offset", 0.0))
            except (KeyError, TypeError, ValueError):
                continue
            self.calibrations[str(hw_id).upper()] = (gain, offset)

    def project_lock_required(self) -> bool:
        """Whether project selection / Open P&ID should be locked right now.

        Locked once the operator has confirmed safe state (the system is
        live) or while any valve is open. Unlocked again once back in a
        not-yet-confirmed, fully-closed state - e.g. freshly connected but
        before Confirm Safe State, or after an abort once everything has
        actually closed.
        """
        return (not self.lockout) or any(self.valve_states.values())

    def close_all_valves(self):
        """Return the system to fully closed without aborting."""
        open_valves = [v for v, is_open in self.valve_states.items() if is_open]
        for v in open_valves:
            self.toggle_valve(v, False)
        if open_valves:
            self.log_event("FULLY_CLOSED", f"Closed: {', '.join(open_valves)}")

    def toggle_throttling(self):
        self.throttling_enabled = not self.throttling_enabled

        status = "ENABLED" if self.throttling_enabled else "DISABLED"
        self.log_event(f"THROTTLING:{status}")

    def toggle_gimbaling(self):
        self.gimbaling_enabled = not self.gimbaling_enabled

        status = "ENABLED" if self.gimbaling_enabled else "DISABLED"
        self.log_event(f"GIMBALING:{status}")
    
    def show_fire_sequence_dialog(self):
        if self.lockout:
            QMessageBox.warning(self.parent, "Abort Active", "Auto fire sequence cannot be activated during an abort")
            return

        # ── Resolve available sequences ──────────────────────────────────────
        named_sequences = getattr(self.project, "sequences", []) if self.project else []
        legacy_steps    = getattr(self.project, "sequence",  []) if self.project else []

        if not named_sequences and not legacy_steps:
            QMessageBox.warning(
                self.parent, "No Sequence",
                "The loaded project has no auto-fire sequence defined.\n"
                "Use the Sequencing tab to build one first."
            )
            return

        # ── Confirmation dialog ──────────────────────────────────────────────
        confirm_dialog = QDialog(self.parent)
        confirm_dialog.setWindowTitle("Confirm Auto-Fire Sequence")
        c_layout = QVBoxLayout(confirm_dialog)

        # Resolve the active sequence steps and name
        steps = legacy_steps
        seq_name = "Legacy Sequence"
        active_id = getattr(self.project, "active_sequence_id", "")
        if named_sequences and active_id:
            for ns in named_sequences:
                if ns.id == active_id:
                    steps = ns.steps
                    seq_name = ns.name
                    break
        elif named_sequences:
            steps = named_sequences[0].steps
            seq_name = named_sequences[0].name

        if not steps:
            QMessageBox.warning(self.parent, "Empty Sequence", "The selected sequence has no steps.")
            return

        info_label = QLabel(
            f"Execute auto-fire sequence: {seq_name}?\n\n"
            f"Steps: {len(steps)}\n"
            f"Total duration: {steps[-1].time_offset:.1f} s"
        )
        c_layout.addWidget(info_label)

        buttons = QDialogButtonBox(QDialogButtonBox.Yes | QDialogButtonBox.Cancel)
        buttons.accepted.connect(confirm_dialog.accept)
        buttons.rejected.connect(confirm_dialog.reject)
        c_layout.addWidget(buttons)
        if confirm_dialog.exec_() != QDialog.Accepted:
            return

        if not any(getattr(s, "is_terminal_count", False) for s in steps):
            QMessageBox.warning(
                self.parent, "No Terminal Count State",
                "The sequence must include a state marked as 'Terminal Count' "
                "before it can be auto-fired.\n\n"
                "Open the Sequencing tab, select the hold state, and check "
                "'This is the Terminal Count state'."
            )
            return

        # ── Schedule each sequence step as a QTimer shot ────────────────────
        # No artificial pre-delay here: the terminal-count hold IS one of the
        # steps below, scheduled at its own time_offset like any other step.
        # A live "TERMINAL COUNT" overlay (see _execute_sequence_step) shows
        # its countdown in real time instead of blocking beforehand - so the
        # operator sees exactly one countdown, not a generic wait stacked on
        # top of the step's own hold.
        self._cancel_sequence()

        for step in steps:
            delay_ms = int(step.time_offset * 1000)
            t = QTimer()
            t.setSingleShot(True)
            t.timeout.connect(lambda s=step: self._execute_sequence_step(s))
            t.start(delay_ms)
            self._sequence_timers.append(t)

        self.log_event("SEQUENCE:START", f"{len(steps)} steps ({seq_name})")


    def _execute_sequence_step(self, step):
        """Apply one sequence step: set every known valve to its desired state.
        
        step.open_valves contains component IDs (cid). These must be translated
        to hardware IDs (hw_id) before calling toggle_valve, which works in hw space.
        """
        if self.lockout:
            return

        # Build a set of hw_ids that should be OPEN for this step
        open_hw_ids: set[str] = set()
        if self.project:
            for cid in step.open_valves:
                comp = self.project.components.get(cid)
                if comp:
                    hw_id = comp.extras.get("hw_id") or comp.label
                    if hw_id:
                        open_hw_ids.add(hw_id)

        # Set every valve in hw space to its desired state
        for hw_id in list(self.valve_states.keys()):
            self.toggle_valve(hw_id, hw_id in open_hw_ids)

        self.ethernet_client.set_system_state(step.name)
        self.signals.system_status.emit(step.name)
        self.log_event("SEQUENCE:STEP", step.name)

        if getattr(step, "is_terminal_count", False):
            tc_length = getattr(self.project.parameters, "terminal_count_s", 10.0)
            self._show_terminal_count_overlay(step.name, tc_length)

    def _show_terminal_count_overlay(self, step_name: str, seconds: float):
        """Live, non-blocking T-minus display for the terminal-count hold.

        This is purely a visual/cancel affordance - it does not delay or
        reschedule anything. The step's own position in the sequence (its
        time_offset relative to the next step) is what actually determines
        how long the hold lasts; this overlay just counts down alongside it.
        """
        length = max(int(round(seconds)), 1)

        dlg = QDialog(self.parent)
        dlg.setWindowTitle("Terminal Count")
        dlg.setMinimumSize(300, 150)
        dlg.setModal(False)
        lay = QVBoxLayout(dlg)

        label = QLabel(f"{step_name}\nT-{length}")
        label.setAlignment(Qt.AlignCenter)
        font = label.font(); font.setPointSize(14); font.setBold(True)
        label.setFont(font)
        lay.addWidget(label)

        cancel_btn = QPushButton("CANCEL SEQUENCE")
        cancel_btn.setStyleSheet("background-color: red; color: white;")
        lay.addWidget(cancel_btn)

        remaining = [length]
        timer = QTimer(dlg)
        timer.setInterval(1000)

        def tick():
            remaining[0] -= 1
            if remaining[0] <= 0:
                timer.stop()
                dlg.close()
            else:
                label.setText(f"{step_name}\nT-{remaining[0]}")

        def do_cancel():
            timer.stop()
            self._cancel_sequence()
            self.close_all_valves()
            self.log_event("SEQUENCE:CANCELED", f"Canceled during terminal count ({step_name})")
            dlg.close()

        timer.timeout.connect(tick)
        cancel_btn.clicked.connect(do_cancel)
        timer.start()
        dlg.show()

    def _cancel_sequence(self):
        """Stop all pending sequence timers."""
        for t in self._sequence_timers:
            t.stop()
        self._sequence_timers.clear()

    def show_abort_countdown_dialog(self, initial_seconds: int, valve_name: str = "UNKNOWN", cmd_type: str = "COMMAND"):
        """Show a non-modal dialog displaying the auto-abort countdown"""
        # Close any existing countdown dialog
        if self.abort_countdown_dialog is not None:
            self.abort_countdown_dialog.close()
            self.abort_countdown_dialog = None
        
        # Create countdown dialog
        self.abort_countdown_dialog = QDialog(self.parent)
        self.abort_countdown_dialog.setWindowTitle("AUTO-ABORT WARNING")
        self.abort_countdown_dialog.setMinimumSize(400, 180)
        countdown_layout = QVBoxLayout(self.abort_countdown_dialog)
        
        # Countdown label
        self.abort_countdown_label = QLabel(f"ABORT IN {initial_seconds} SECONDS")
        self.abort_countdown_label.setAlignment(Qt.AlignCenter)
        font = self.abort_countdown_label.font()
        font.setPointSize(16)
        font.setBold(True)
        self.abort_countdown_label.setFont(font)
        countdown_layout.addWidget(self.abort_countdown_label)
        
        # Warning message with valve details
        warning_label = QLabel(f"Valve {valve_name} {cmd_type} command failed!\nClick button below to cancel abort.")
        warning_label.setAlignment(Qt.AlignCenter)
        countdown_layout.addWidget(warning_label)
        
        # Cancel button
        cancel_btn = QPushButton("CONFIRM SAFE STATE\n(Cancel Abort)")
        cancel_btn.setStyleSheet("background-color: green; color: white; font-size: 14px; padding: 10px;")
        cancel_btn.clicked.connect(self._cancel_abort_countdown)
        countdown_layout.addWidget(cancel_btn)
        
        # Show dialog non-modally
        self.abort_countdown_dialog.show()
    
    def update_abort_countdown_dialog(self, remaining_seconds: int):
        """Update the countdown display"""
        if self.abort_countdown_dialog is not None and self.abort_countdown_label is not None:
            self.abort_countdown_label.setText(f"ABORT IN {remaining_seconds} SECONDS")
    
    def close_abort_countdown_dialog(self):
        """Close the abort countdown dialog"""
        if self.abort_countdown_dialog is not None:
            self.abort_countdown_dialog.close()
            self.abort_countdown_dialog = None
            self.abort_countdown_label = None
    
    def _cancel_abort_countdown(self):
        """Called when user clicks the cancel button in the countdown dialog"""
        self.confirm_safe_state()
        self.close_abort_countdown_dialog()

    def update_valve_state(self, valve_name: str, new_val: str):        
        # Only update the internal state if it was confirmed open or closed
        if new_val == "OPEN":
            self.valve_states[valve_name] = True
        elif new_val == "CLOSED":
            self.valve_states[valve_name] = False

        # Update the manual valve buttons
        if valve_name in self.manual_valve_buttons:
            if new_val == "OPEN":
                self.manual_valve_buttons[valve_name].setStyleSheet(f"background-color: green; color: white;")
            elif new_val == "CLOSED":
                self.manual_valve_buttons[valve_name].setStyleSheet(f"background-color: red; color: white;")
            elif new_val == "PENDING":
                self.manual_valve_buttons[valve_name].setStyleSheet(f"background-color: gray; color: white;")

    def show_manual_valve_control(self):
        if self.lockout:
            QMessageBox.warning(self.parent, "Lockout Active", "Manual control is disabled during abort")
            return
        
        # Close existing dialog if open
        if self.manual_valve_dialog:
            self.manual_valve_dialog.close()
        
        dialog = QDialog(self.parent)
        dialog.setWindowTitle("Manual Valve Control")
        dialog.setModal(False)  # Allow interaction with main window
        layout = QVBoxLayout(dialog)
        
        # Store reference to dialog
        self.manual_valve_dialog = dialog
        
        # Get actual valve names from diagram
        valve_names = list(self.valve_states.keys())
        
        # Create buttons and store references
        self.manual_valve_buttons = {}
        for valve in valve_names:
            current_state = self.valve_states[valve]
            btn = QPushButton(valve)
            color = "green" if current_state else "red"
            btn.setStyleSheet(f"background-color: {color}; color: white;")
            # Store button reference
            self.manual_valve_buttons[valve] = btn
            # Connect click handler with valve name
            btn.clicked.connect(lambda checked, v=valve: self.toggle_valve(v))
            layout.addWidget(btn)
            
        dialog.show()

    def toggle_valve(self, valve_name: str, state: bool=None):
        if self.lockout:
            return

        if state is None:
            new_state = not self.valve_states[valve_name]
        else:
            new_state = state

        # Only send command if valve state is actually changing
        if valve_name in self.valve_states and self.valve_states[valve_name] == new_state:
            # Valve already in desired state, no need to send command
            return

        # self.valve_states[valve_name] = new_state     # needs to update when we get a response
        self.signals.valve_updated.emit(valve_name, "PENDING")
        
        # Update manual valve button if dialog is open
        # if valve_name in self.manual_valve_buttons:
        #     # color = "green" if new_state else "red"
        #     self.manual_valve_buttons[valve_name].setStyleSheet(
        #         f"background-color: gray; color: white;"
        #     )
        
        try:
            self.ethernet_client.send_valve_command(valve_name, new_state)
        except Exception:
            pass

        self.log_event("VALVE_CHANGED", f"{valve_name}:{new_state}")

        self.pending_valve_commands[valve_name] = {
            "state": new_state,
            "last_sent": QDateTime.currentMSecsSinceEpoch(),
            "retries": 0
        }

    def check_valve_command_timeouts(self):
        current_time = QDateTime.currentMSecsSinceEpoch()
        valves_to_delete = []

        for valve_name, info in self.pending_valve_commands.items():
            if current_time - info["last_sent"] > self.valve_retry_timeout:
                if info["retries"] < self.valve_max_retries:

                    info["retries"] += 1
                    info["last_sent"] = current_time
                    try:
                        self.ethernet_client.send_valve_command(valve_name, info["state"])
                        self.log_event("VALVE_RETRY", f"{valve_name}:{info['state']} (Attempt {info['retries']})")
                    except Exception:
                        pass
                else:
                    print(f"Timeout: Max retries reached for valve {valve_name}")
                    self.log_event("VALVE_TIMEOUT", f"{valve_name}:{info['state']}")
                    valves_to_delete.append(valve_name)

        for valve_name in valves_to_delete:
            del self.pending_valve_commands[valve_name]


    def apply_operation(self, operation: str):
        if self.lockout:
            return

        self.ethernet_client.set_system_state(operation)
        self.signals.system_status.emit(operation)
        self.log_event("OPERATION", operation)

        if self.project and getattr(self.project, "sequence", None):
            step = next(
                (s for s in self.project.sequence if s.name == operation),
                None
            )
            if step is not None:
                open_hw_ids: set[str] = set()
                for cid in step.open_valves:
                    comp = self.project.components.get(cid)
                    if comp:
                        hw_id = comp.extras.get("hw_id") or comp.label
                        if hw_id:
                            open_hw_ids.add(hw_id)
                for hw_id in list(self.valve_states.keys()):
                    self.toggle_valve(hw_id, hw_id in open_hw_ids)
                return

        self.log_event("OPERATION:WARN", f"No sequence step named '{operation}' in project")